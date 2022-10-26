import base64
import json
import openpyxl
from global_args import *
from core.base.logger import logger
from core.base.md5_encryption import *


class DataFile(object):
    def __init__(self, filename=None, sheet=None):
        self.filename = os.path.join(main_path, 'data_file', filename)
        self.sheet = sheet
        self.workbook, self.worksheet = self.open_file()
        self.max_row = self.worksheet.max_row
        self.max_column = self.worksheet.max_column

    def open_file(self):
        """打开参数文件"""
        logger.info(f'解析文件：{self.filename}')
        try:
            workbook = openpyxl.load_workbook(self.filename)
            worksheet = workbook[self.sheet] if self.sheet else workbook.active
        except FileNotFoundError:
            logger.error('请输入正确的文件名！！！')
            assert 0
        except KeyError:
            logger.error('请输入正确的sheet名！！！')
            assert 0
        return workbook, worksheet

    def close_file(self):
        """关闭文件"""
        logger.info(f'关闭文件：{self.filename}')
        if self.workbook:
            self.workbook.close()

    def parse_api(self, row=1):
        """获取api"""
        return self.worksheet.cell(row=row).value

    def parse_file(self):
        """解析文件，获取参数"""
        row_list = []
        for r in range(self.max_row - 1):  # 遍历获取每一行数据
            arg_list = []
            for each in self.worksheet.iter_cols(min_row=2):  # 从第二行开始
                arg_list.append(each[r].value)
            row_list.append(arg_list)
        self.data_clean(row_list)
        return row_list

    def data_clean(self, row_list):
        """数据格式化"""
        for row in row_list:
            # 检测接口是否'/'开头
            row[column_dict['api'] - 1] = self.api_clean(row[column_dict['api'] - 1])
            # 数据-字符转json
            row[column_dict['json'] - 1] = self.params_clean(row[column_dict['json'] - 1])
            # 添加验证
            row = self.verify_clean(row)

    @staticmethod
    def api_clean(val):
        """检测api格式"""
        if val and val[0] != '/':
            val = f'/{val}'
        return val

    @staticmethod
    def params_clean(val):
        """参数处理"""
        if val:
            val = json.loads(val)
            if 'filemd5' in val:
                val['filemd5'] = md5_encode(base64.b64encode(val['filename'].encode()))
        return val

    @staticmethod
    def verify_clean(row):
        verify = row[column_dict['verify'] - 1]
        json_ = row[column_dict['json'] - 1]
        if not verify:
            return row
        arg_list = verify.split('+')
        arg_list = [arg.strip() for arg in arg_list]
        verify_str = ''
        for arg in arg_list:
            if arg.lower() == 'app_token':
                verify_str += app_token
            elif arg not in json_:
                logger.error(f'verify参数--{arg} 不在json数据内')
                return False
            else:
                verify_str += str(json_[arg])
        if not row[column_dict['json'] - 1]:  # get请求可能只存在verify参数
            row[column_dict['json'] - 1] = {}
        row[column_dict['json'] - 1]['verify'] = md5_encode(verify_str)
        return row
